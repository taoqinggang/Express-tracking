from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.db import IntegrityError
from .forms import DeliveryForm
from django.shortcuts import render, redirect
from .models import Delivery
from django.contrib.auth.mixins import LoginRequiredMixin
import requests
import json
from django.conf import settings
from django.http import JsonResponse
from .models import DailyQueryCount
from django.views.generic import ListView
import os
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from openpyxl import load_workbook
from django.http import FileResponse



@csrf_exempt
def import_excel(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]
        try:
            wb = load_workbook(excel_file)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                username, phone, address, sent_tracking_number, return_tracking_number = row
                Delivery.objects.create(
                    username=username,
                    phone=phone,
                    address=address,
                    sent_tracking_number=sent_tracking_number,
                    return_tracking_number=return_tracking_number,
                    fill_date=now(),
                )
            return JsonResponse({"status": "success", "message": "导入成功"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})
    return JsonResponse({"status": "error", "message": "无效的请求"})

def download_sample(request):
    file_path = os.path.join(settings.STATICFILES_DIRS[0], 'sample.xlsx')
    if not os.path.exists(file_path):
        return JsonResponse({"status": "error", "message": "样例文件不存在"})
    response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename='sample.xlsx')
    return response



# 批量删除
def batch_delete(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            if not ids:
                return JsonResponse({'status': 'error', 'message': '未选择任何记录'})

            # 删除对应的记录
            Delivery.objects.filter(id__in=ids).delete()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    else:
        return JsonResponse({'status': 'error', 'message': '仅支持 POST 请求'})




def search_by_phone(request):
    phone = request.GET.get('phone')
    try:
        delivery = Delivery.objects.get(phone=phone)
        return JsonResponse({
            'status': 'success',
            'delivery': {
                'id': delivery.id,
                'username': delivery.username,
                'fill_date': delivery.fill_date.strftime('%Y-%m-%d %H:%M'),
                'phone': delivery.phone,
                'address': delivery.address,
                'sent_tracking_number': delivery.sent_tracking_number,
                'return_tracking_number': delivery.return_tracking_number,
            }
        })
    except Delivery.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': '未找到对应的用户'})





def track_delivery(request):
    # 检查当天的查询次数是否达到上限
    if not DailyQueryCount.increment_count():
        return JsonResponse({'status': 'error', 'message': '今日查询次数已达上限'}, status=403)

    if request.method == "GET" and "tracking_number" in request.GET:
        tracking_number = request.GET.get("tracking_number")
        url = "https://alicloudmarket8002.kdniao.com/api/track/8002"
        headers = {
            "Authorization": f"APPCODE {settings.KDNIAO_APPCODE}",
            "Content-Type": "application/json"
        }
        payload = json.dumps({
            "LogisticCode": tracking_number
        })

        try:
            response = requests.post(url, headers=headers, data=payload, verify=False)
            response_data = response.json()
            if response.status_code == 200 and response_data.get("Success"):
                return JsonResponse({
                    "status": "success",
                    "data": response_data["Traces"]
                })
            else:
                return JsonResponse({
                    "status": "error",
                    "message": response_data.get("Reason", "查询失败")
                })
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "message": str(e)
            })
    else:
        return JsonResponse({
            "status": "error",
            "message": "无效的请求"
        })

class DeliveryListView(LoginRequiredMixin, ListView):
    model = Delivery
    context_object_name = 'deliveries'
    template_name = 'delivery/delivery_list.html'
    login_url = 'login'  # 设置未登录时的跳转页面
    # redirect_field_name = 'redirect_to'
    paginate_by = 7  # 每页显示 8 条数据
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['can_delete'] = self.request.user.has_perm('delivery.delete_delivery')
        return context
    def get_queryset(self):
        return Delivery.objects.all().order_by('-fill_date')  # 按填报时间降序排序



# 创建视图：添加新的寄修记录
class DeliveryCreateView(CreateView):
    model = Delivery
    form_class = DeliveryForm
    template_name = 'delivery/delivery_form.html'
    success_url = reverse_lazy('delivery-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action_text'] = '编辑'
        return context

# 更新视图：编辑已存在的寄修记录
class DeliveryUpdateView(UpdateView):
    model = Delivery
    form_class = DeliveryForm
    template_name = 'delivery/delivery_form.html'
    success_url = reverse_lazy('delivery-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action_text'] = '编辑'
        return context





# 删除视图：删除寄修记录
class DeliveryDeleteView(DeleteView):
    model = Delivery
    template_name = 'delivery/delivery_confirm_delete.html'
    success_url = reverse_lazy('delivery-list')




def login_view(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('delivery-list')  # 登录成功后跳转到首页
        else:
            messages.error(request, "用户名或密码错误")
    return render(request, 'delivery/login.html')

def add_delivery(request):
    if request.method == 'POST':
        form = DeliveryForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('delivery-list')
            except IntegrityError:
                # 如果手机号重复，传递错误消息到模板
                return render(request, 'delivery/delivery_form.html', {
                    'form': form,
                    'error_message': "手机号已经存在，不能重复添加。"
                })
    else:
        form = DeliveryForm()
    return render(request, 'delivery/delivery_form.html', {'form': form})


